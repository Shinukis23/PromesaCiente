# Programa para calcular Los trabajos subidos a Produccion por Vendedor, Due-date calculado, 
# Diferencia Due-date Calculado vs. Due-Date en Sistema de Produccion
# Junio 5/ 2023
# Modificado Agosto 4/2023   s
# Modificado Agosto 8/2023 se modifico instruccion para eliminar filas con datos nulos en la columna Diferencia DueDates
# Modificado Agosto 9/2023 se modifico instruccion para eliminar filas de las tiendas 20 y 21

import os
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook

# Get the current working directory
directory = os.getcwd()
plt.rc('figure',figsize = (70,20))
# Initialize an empty list to hold the dataframes from all .xlsx files
dfs = []
#####

print(pd.__version__)
# Supongamos que tienes un DataFrame llamado "dataframe" con tus datos

# Ruta del archivo de Excel donde quieres agregar el DataFrame
####archivo_excel = 'Reporte_Diario_Due_Date_Mes_dia.xlsx'
# Nombre de la nueva hoja donde agregarás el DataFrame
nombre_nueva_hoja = 'Due_Date'

# Crear un objeto ExcelWriter para cargar el archivo de Excel
#writer = pd.ExcelWriter(archivo_excel, engine='openpyxl')
####writer = pd.ExcelWriter(archivo_excel, mode='w',engine='xlsxwriter')
###libro_existente = openpyxl.load_workbook(archivo_existente)
###libro_nuevo = openpyxl.Workbook()
####workbook = writer.book
####writer.sheets={nombre_nueva_hoja:workbook.add_worksheet()}
#load_workbook(archivo_excel, read_only=True)

#writer.book = load_workbook(archivo_excel)

## Si la hoja ya existe, borrarla para evitar conflictos
#try:
#    writer.book.remove(writer.book[nombre_nueva_hoja])
#except KeyError:
#    pass

####print('aqui')



###
datos = pd.read_excel(r'DuedateRutas_Reporte.xlsx',sheet_name='Semanal')
#datos = pd.read_csv()
datos.info()
indexDeleted = datos[(datos['Job Status'].isin(['Voided', 'New'])) | (datos['Diferencia DueDates'].isnull()) | (datos['Part Store #'].isin([20,21]))].index
datos.drop(indexDeleted,inplace=True) 
datos.to_excel(r'Revision.xlsx')

datos['Menor que 0'] = datos['Diferencia DueDates'] < 0
"""def crear_lista3(group):
	#return list(zip(group['Created'], group['Menor que 0'],group['Due_Date_Vendedor']))
	count_true = group['Menor que 0'].sum()  # Cuenta cuántos son True
	count_false = len(group) - count_true    # Cuenta cuántos son False
	#return [(count_true, count_false)]
	return [(list(zip(group['Due_Date_Vendedor'], group['Menor que 0'])),count_true, count_false)]
# Agrupa los datos por 'Salesperson', 'Customer' y 'Created' con una frecuencia de 1 minuto
result3 = datos.groupby([
    'Created by (Salesperson)',
    'Customer',
    pd.Grouper(key='Created', freq='1T'),
    pd.Grouper(key='Due_Date_Calculado') # Agrupa por hora con diferencia de 1 minuto
]).apply(crear_lista3)

result3.to_excel(r'DueDate_Grupo7.xlsx')
"""
#Perfecto
def crear_lista(group):
    count_true = group['Menor que 0'].sum()  # Cuenta cuántos son True
    count_false = len(group) - count_true    # Cuenta cuántos son False

    due_dates = group['Due_Date_Vendedor'].tolist()
    
    return [(count_true, count_false, due_dates)]

    #return [(count_true, count_false)]

result = datos.groupby([
    'Created by (Salesperson)',
    'Customer',
    pd.Grouper(key='Created', freq='1T'),
    pd.Grouper(key='Due_Date_Calculado')
]).apply(crear_lista)

result.to_excel(r'DueDate_Grupo4.xlsx')

#############
def crear_lista2(group):
    count_true = group['Menor que 0'].sum()  # Cuenta cuántos son True
    count_false = len(group) - count_true    # Cuenta cuántos son False
    return [(count_true, count_false)]

result2 = datos.groupby([
    'Created by (Salesperson)',
    'Customer',
    pd.Grouper(key='Created', freq='1T'),
    pd.Grouper(key='Due_Date_Calculado')
]).apply(crear_lista2)


# Agregar filas con la cantidad de True y False por Salesperson
salesperson_counts = datos.groupby('Created by (Salesperson)')['Menor que 0'].value_counts().unstack(fill_value=0)
salesperson_counts.reset_index(inplace=True)
salesperson_counts.rename(columns={True: 'True Salesperson', False: 'False Salesperson'}, inplace=True)

# Convertir la Serie 'result' en un DataFrame
result2 = result2.to_frame(name='Counts')

# Fusionar el DataFrame 'result' con 'salesperson_counts'
result2 = result2.merge(salesperson_counts, how='left', left_on='Created by (Salesperson)', right_on='Created by (Salesperson)')


result2.to_excel(r'DueDate_Grupo6.xlsx')
#result1.to_excel(r'porgruponew1.xlsx')

#""" Este tambien esta bien
def crear_lista3(group):
    count_true = group['Menor que 0'].sum()  # Cuenta cuántos son True
    count_false = len(group) - count_true    # Cuenta cuántos son False
    
    rows = [
        (group.name[0], group.name[1], group.name[2], count_true, True),
        (group.name[0], group.name[1], group.name[2], count_false, False)
    ]
    
    return pd.DataFrame(rows, columns=['Created by (Salesperson)', 'Customer', 'Created', 'Count', 'Menor que 0'])

# Agrupa los datos por 'Salesperson', 'Customer' y 'Created' con una frecuencia de 1 minuto
result3 = datos.groupby([
    'Created by (Salesperson)',
    'Customer',
    pd.Grouper(key='Created', freq='1T'),
    pd.Grouper(key='Due_Date_Calculado')  # Agrupa por hora con diferencia de 1 minuto
]).apply(crear_lista3)

# Resetear el índice del DataFrame
result3.reset_index(drop=True, inplace=True)

# Filtrar las filas con True
true_rows = result3[result3['Menor que 0']]
false_rows = result3[result3['Menor que 0']==0]

# Calcular la cantidad de True por Salesperson
salesperson_counts = true_rows.groupby('Created by (Salesperson)')['Count'].sum().reset_index()
salesperson_countsfalse = false_rows.groupby('Created by (Salesperson)')['Count'].sum().reset_index()
# Agregar las filas de True Count al DataFrame original
result3 = pd.concat([result3, salesperson_counts,salesperson_countsfalse], ignore_index=True, sort=False)

# Ordenar el DataFrame por Salesperson y Created
result3.sort_values(by=['Created by (Salesperson)', 'Created'], inplace=True)

# Guardar el DataFrame en un archivo Excel
result3.to_excel(r'DueDate_Grupo8.xlsx', index=False)

resultor = datos.groupby(['Created by (Salesperson)', 'Menor que 0']).size().reset_index(name='Count')


#print(result)

# Guardar el DataFrame en la nueva hoja
#result.to_excel(writer, sheet_name=nombre_nueva_hoja, index=False)
####result.to_excel(writer, sheet_name=nombre_nueva_hoja, index=False)

# Guardar los cambios en el archivo de Excel
####writer.save()
####writer.close()
resultor.to_excel(r'DueDate_Grupo.xlsx')

print(vendedores)
#vendedores.groups()
fig= vendedores['Diferencia DueDates'].value_counts().plot.bar(color= 'blue')
fig.set_ylabel('Cantidad')
fig.set_title('Diferencia de dias DueDate Vendedor-Calculado',{'fontsize':6})
dl = pds.DataFrame()

# Concatenate all dataframes in the dfs list into a single dataframe
concatenated_data = pd.concat(dfs, ignore_index=True)
concatenated_data.drop_duplicates(subset='Job #',keep='last', inplace=True)
concatenated_data.to_excel(r'ReporteProduccionDB.xlsx', index=False)

#Job #   Job Status  Reason  Due
#Drop Location   R # Stock # Interchange Part Description Summary    Part Price  Created Pull Started    Pull Started By Pulled Finished Pulled Finish By    Ship Via    Inspector   Order Store #   Part Store #
#rutas= pd.read_excel(r'Rutas pendientes.xls')
#merged_data = pd.merge(rutas,concatenated_data[['Job #','Drop Location','R #','Stock #','Interchange','Part Description Summary',
#    'Part Price','Created','Ship Via','Order Store #','Part Store #','Due']],on=['Job #'],how="left")
# Write the concatenated data to a new .xlsx file
#merged_data["Delivery time"]= datetime.now() 
#indexDeleted = ds2[ds2['Job Status'] ==  'Pickup'].index
#ds2.drop(indexDeleted,inplace=True)
#indexDeleted = merged_data[merged_data['Drop Location'] == ' '].index  # dejando solo las 253 en copia de audit trial
#merged_data.drop(indexDeleted,inplace=True)
#merged_data=merged_data.dropna(subset=['Drop Location'])
output_file_path = os.path.join(directory, 'ReporteProduccionDB.xlsx')
concatenated_data.to_excel(output_file_path, index=False)
